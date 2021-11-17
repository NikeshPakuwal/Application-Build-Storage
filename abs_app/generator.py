from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
from django.core.files.base import ContentFile
from django.core.files.storage import FileSystemStorage
from .models import Version

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


def generate_version_pdf(request):
    version = Version.objects.all()

    html_string = render_to_string('version_pdf.html', {'version': version})
    pdf = HTML(string=html_string).write_pdf()

    c = ContentFile(pdf)
    fs = FileSystemStorage()
    filename = fs.save("version-list.pdf", c)
    uploaded_file_url = fs.url(filename)
    return uploaded_file_url


def get_version_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Version List"

    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    fill = PatternFill(
        fill_type=None,
        start_color='00FFCC00',
        end_color='00FFCC00'
    )

    alignment = Alignment(
        horizontal='left',
        vertical='center',
        shrink_to_fit=True
    )

    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        shrink_to_fit=True
    )

    for i in range(3):
        ws.row_dimensions[i].height = 25

    ws.merge_cells('A1:F1')
    ws.merge_cells('A2:F2')
    ws.merge_cells('A3:F3')
    ws.merge_cells('A4:F4')
    ws.merge_cells('A5:F5')

    h1 = ws['A1']
    h2 = ws['A2']
    h1.font, h2.font = [
        Font(bold=True, size=12, color='800080'),
        Font(bold=True, size=12, color='800080'),
        
    ]

    h1.value = "List of Application Version"
    h2.value = "Available Version"
   
    h1.alignment, h2.alignment = [
        center_alignment, center_alignment
    ]

    start_row = 7

    others_width = 10
    row = 12

    ws.column_dimensions[get_column_letter(2)].width = 35
    ws.column_dimensions[get_column_letter(1)].width = 5
    ws.column_dimensions[get_column_letter(3)].width = 25
    ws.column_dimensions[get_column_letter(4)].width = 20
    ws.column_dimensions[get_column_letter(5)].width = 70

    version = Version.objects.all()

    ws.cell(start_row, 1, "S.N")
    ws.cell(start_row, 1).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(start_row, 1).alignment = alignment
    ws.cell(start_row, 1).border = border
    ws.cell(start_row, 1).fill = PatternFill(
        fill_type='solid', start_color='808080', end_color='808080')

    ws.cell(start_row, 2, "Title")
    ws.cell(start_row, 2).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(start_row, 2).alignment = alignment
    ws.cell(start_row, 2).border = border
    ws.cell(start_row, 2).fill = PatternFill(
        fill_type='solid', start_color='808080', end_color='808080')

    ws.cell(start_row, 3, "Build Name")
    ws.cell(start_row, 3).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(start_row, 3).alignment = alignment
    ws.cell(start_row, 3).border = border
    ws.cell(start_row, 3).fill = PatternFill(
        fill_type='solid', start_color='808080', end_color='808080')

    ws.cell(start_row, 4, "Version")
    ws.cell(start_row, 4).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(start_row, 4).alignment = alignment
    ws.cell(start_row, 4).border = border
    ws.cell(start_row, 4).fill = PatternFill(
        fill_type='solid', start_color='808080', end_color='808080')

    ws.cell(start_row, 5, "Build Notes")
    ws.cell(start_row, 5).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(start_row, 5).alignment = alignment
    ws.cell(start_row, 5).border = border
    ws.cell(start_row, 5).fill = PatternFill(
        fill_type='solid', start_color='808080', end_color='808080')

    row = start_row + 1
    for i in range(len(version)):
        ws.cell(i+row, 1, i+1)
        ws.cell(i+row, 1).alignment = alignment
        ws.cell(i+row, 1).border = border
        ws.cell(i+row, 1).fill = fill

        ws.cell(i+row, 2, version[i].title)
        ws.cell(i+row, 2).alignment = alignment
        ws.cell(i+row, 2).border = border
        ws.cell(i+row, 2).fill = fill

        ws.cell(i+row, 3, version[i].build_name)
        ws.cell(i+row, 3).alignment = alignment
        ws.cell(i+row, 3).border = border
        ws.cell(i+row, 3).fill = fill

        ws.cell(i+row, 4, version[i].version)
        ws.cell(i+row, 4).alignment = alignment
        ws.cell(i+row, 4).border = border
        ws.cell(i+row, 4).fill = fill

        ws.cell(i+row, 5, version[i].build_notes)
        ws.cell(i+row, 5).alignment = alignment
        ws.cell(i+row, 5).border = border
        ws.cell(i+row, 5).fill = fill

    path = f"media/version-list.xlsx"
    wb.save(path)
    return path